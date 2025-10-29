from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import io
from fastapi.responses import StreamingResponse, HTMLResponse
from fastapi import APIRouter, Depends, Request
from sqlalchemy.orm import Session
from sqlalchemy import func, and_
from app.database import get_db
from fastapi.templating import Jinja2Templates
from app import models
from app.utils.timezone import formatear_hora_panama

router = APIRouter()
templates = Jinja2Templates(directory="app/templates")


# 🧭 Vista principal del tablero
@router.get("/tablero", response_class=HTMLResponse)
async def mostrar_tablero(request: Request, db: Session = Depends(get_db)):
    tablero = {  # Estado → listado de placas
        "Patio": [],
        "Bodega": [],
        "Cargando": [],
    }

    punto_a_estado = {
        "punto1": "Patio",
        "punto2": "Bodega",
        "punto3": "Cargando",
        "punto4": "Cargando",
        "punto5": "Cargando",
    }

    # Subconsulta para obtener la última marca de tiempo por ciclo
    ultimos_por_ciclo = (
        db.query(
            models.Escaneo.ciclo_id.label("ciclo_id"),
            func.max(models.Escaneo.fecha_hora).label("ultima_fecha")
        )
        .group_by(models.Escaneo.ciclo_id)
        .subquery()
    )

    # Último escaneo por ciclo con su punto asociado
    ultimo_escaneo = (
        db.query(
            models.Escaneo.ciclo_id.label("ciclo_id"),
            models.Escaneo.punto.label("punto"),
            models.Escaneo.fecha_hora.label("fecha_hora")
        )
        .join(
            ultimos_por_ciclo,
            and_(
                models.Escaneo.ciclo_id == ultimos_por_ciclo.c.ciclo_id,
                models.Escaneo.fecha_hora == ultimos_por_ciclo.c.ultima_fecha
            )
        )
        .subquery()
    )

    registros = (
        db.query(
            models.Sesion.placa,
            ultimo_escaneo.c.punto,
            ultimo_escaneo.c.fecha_hora
        )
        .join(models.Ciclo, models.Ciclo.sesion_id == models.Sesion.id)
        .join(ultimo_escaneo, ultimo_escaneo.c.ciclo_id == models.Ciclo.id)
        .filter(
            models.Ciclo.completado == False,
            models.Sesion.cerrada == False
        )
        .order_by(ultimo_escaneo.c.fecha_hora.desc())
        .all()
    )

    for placa, punto, fecha in registros:
        estado = punto_a_estado.get(punto)
        if not estado:
            continue
        tablero[estado].append({
            "placa": placa,
            "hora": formatear_hora_panama(fecha),
        })

    return templates.TemplateResponse("tablero.html", {
        "request": request,
        "tablero": tablero,
    })


# 📊 Descarga de informes Excel
@router.get("/descargar_informe")
def descargar_informe(fechaInicio: str, fechaFin: str, db: Session = Depends(get_db)):
    inicio = datetime.strptime(fechaInicio, "%Y-%m-%d")
    fin = datetime.strptime(fechaFin, "%Y-%m-%d")

    # 🔹 Datos simulados (luego obtendrás esto desde tu BD)
    ciclos = [
        {"placa": "HE2345", "inicio": "2025-10-25 08:00", "fin": "2025-10-25 09:30", "saltos": False, "completado": True},
        {"placa": "HE7865", "inicio": "2025-10-25 09:00", "fin": "2025-10-25 10:10", "saltos": True, "completado": True},
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Informe QRLogix"

    # 🧭 Cabeceras
    headers = ["Placa", "Inicio", "Fin", "Duración (min)", "Saltos", "Estado"]
    ws.append(headers)

    # 🎨 Estilos base
    header_font = Font(bold=True, color="FFFFFF", size=12, name="Arial")
    header_fill = PatternFill("solid", fgColor="071D49")  # Azul Argos
    center_align = Alignment(horizontal="center", vertical="center")
    normal_font = Font(color="071D49", size=11, name="Arial")

    # 🖌️ Estilo de cabecera
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # 🧾 Agregar datos
    for c in ciclos:
        inicio_dt = datetime.strptime(c["inicio"], "%Y-%m-%d %H:%M")
        fin_dt = datetime.strptime(c["fin"], "%Y-%m-%d %H:%M")
        duracion = round((fin_dt - inicio_dt).total_seconds() / 60, 1)
        ws.append([
            c["placa"],
            c["inicio"],
            c["fin"],
            duracion,
            "Sí" if c["saltos"] else "No",
            "Completado" if c["completado"] else "Incompleto"
        ])

    # 📏 Ajuste de columnas
    column_widths = [14, 20, 20, 18, 10, 15]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = width

    # 🧱 Bordes suaves
    thin = Side(border_style="thin", color="E0E0E0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=6):
        for cell in row:
            cell.border = border
            cell.font = normal_font
            cell.alignment = center_align

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:F1"

    # 💾 Guardar en flujo para enviar como descarga
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = f"informe_qrlogix_{fechaInicio}_a_{fechaFin}.xlsx"
    headers = {"Content-Disposition": f"attachment; filename={filename}"}

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers
    )
